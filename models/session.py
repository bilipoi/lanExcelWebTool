"""
文件会话模型
"""
import os
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import DATA_DIR, SNAP_DIR, USER_COLORS, MAX_SNAPSHOTS
from utils import rel_path, calc_data_hash, get_default_data
from services.style_service import load_styles
from services.type_service import load_types


class FileSession:
    """单个文件的会话状态"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath                    # 文件绝对路径
        self.rel_path = rel_path(filepath)          # 相对路径（作为 room name）
        self.spreadsheet_data: Dict[str, List[List[str]]] = {}
        self.cell_styles: Dict[str, Any] = {}       # 单元格样式
        self.cell_types: Dict[str, Any] = {}        # 列类型配置
        self.last_saved_hash: str = ''              # 上次保存的数据指纹
        # 用户信息：{sid: {'username': str, 'color': str, 'selection': {sheet, row, col}}}
        self.online_users: Dict[str, Dict[str, Any]] = {}
        self.readonly: bool = False                 # 只读状态
        self._color_index = 0                       # 颜色分配索引
        self._load_from_disk()
    
    def _load_from_disk(self) -> None:
        """从磁盘加载 Excel 数据 - 优化版本"""
        if not os.path.exists(self.filepath):
            # 文件不存在，创建空白
            self.spreadsheet_data = {'Sheet1': get_default_data()}
            self.last_saved_hash = calc_data_hash(self.spreadsheet_data)
            return
        
        wb = load_workbook(self.filepath, data_only=True, read_only=False)
        total_cells = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not isinstance(ws, Worksheet):
                continue
            
            # 使用 iter_rows 批量读取，比逐 cell 访问快 10-100 倍
            rows: List[List[str]] = []
            
            # 获取实际数据范围
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 如果没有数据，创建最小空表
            if max_row == 0 or max_col == 0:
                rows = [['' for _ in range(26)] for _ in range(50)]
            else:
                # 限制最大加载范围，避免超大文件拖慢系统
                max_row = min(max_row, 1000)  # 最多加载 1000 行
                max_col = min(max_col, 52)    # 最多加载 52 列 (A-Z, AA-AZ)
                
                # 使用 iter_rows 批量读取
                for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                    row_data = [str(val) if val is not None else '' for val in row]
                    rows.append(row_data)
                    total_cells += len(row_data)
                
                # 确保至少有 50 行，保持界面一致性
                while len(rows) < 50:
                    rows.append(['' for _ in range(max(26, max_col))])
            
            self.spreadsheet_data[sheet_name] = rows
        
        wb.close()
        self.last_saved_hash = calc_data_hash(self.spreadsheet_data)
        print(f"[性能] 加载完成: {len(self.spreadsheet_data)} 个 Sheet, 共 {total_cells} 个单元格")
        
        # 加载只读权限
        from services.meta_service import get_file_meta
        m = get_file_meta(self.rel_path)
        self.readonly = m.get('readonly', False)
        
        # 加载样式
        self.cell_styles = load_styles(self.filepath)
        
        # 加载列类型
        self.cell_types = load_types(self.filepath)
    
    def save(self) -> tuple[bool, Optional[str]]:
        """保存到磁盘，如有变更则返回 (True, None)，否则返回 (False, None)
        如果保存失败，返回 (False, error_message)
        """
        current_hash = calc_data_hash(self.spreadsheet_data)
        if current_hash == self.last_saved_hash:
            return False, None  # 无变更
        
        try:
            wb = Workbook()
            default_sheet = wb.active
            if default_sheet is not None:
                wb.remove(default_sheet)
            
            for sheet_name, data in self.spreadsheet_data.items():
                ws = wb.create_sheet(title=sheet_name)
                # 使用 append 批量写入，比逐 cell 快 5-10 倍
                for row in data:
                    # 过滤掉空行，减少文件大小
                    processed_row = []
                    for val in row:
                        if val and val.strip():
                            try:
                                # 尝试转换为数字
                                processed_row.append(float(val))
                            except (ValueError, TypeError):
                                processed_row.append(val)
                        else:
                            processed_row.append(None)  # 空值用 None，Excel 会优化存储
                    ws.append(processed_row)
            
            wb.save(self.filepath)
            wb.close()
            
            self.last_saved_hash = current_hash
            return True, None
        except Exception as e:
            error_msg = str(e)
            print(f"保存文件失败: {error_msg}")
            return False, error_msg
    
    def _get_next_color(self) -> str:
        """获取下一个可用颜色"""
        color = USER_COLORS[self._color_index % len(USER_COLORS)]
        self._color_index += 1
        return color
    
    def add_user(self, sid: str, username: str) -> str:
        """添加用户，返回分配的颜色"""
        color = self._get_next_color()
        self.online_users[sid] = {
            'username': username,
            'color': color,
            'selection': None  # {sheet, row, col}
        }
        return color
    
    def remove_user(self, sid: str) -> Optional[Dict[str, Any]]:
        """移除用户"""
        user_info = self.online_users.pop(sid, None)
        return user_info
    
    def update_selection(self, sid: str, sheet: str, row: int, col: int) -> Optional[Dict[str, Any]]:
        """更新用户选中的单元格"""
        if sid not in self.online_users:
            return None
        self.online_users[sid]['selection'] = {'sheet': sheet, 'row': row, 'col': col}
        return self.online_users[sid]
    
    def clear_selection(self, sid: str) -> Optional[Dict[str, Any]]:
        """清除用户选中"""
        if sid not in self.online_users:
            return None
        self.online_users[sid]['selection'] = None
        return self.online_users[sid]
    
    def get_user_list(self) -> List[Dict[str, Any]]:
        """获取用户列表（包含颜色和选中信息）"""
        return [
            {
                'username': info['username'],
                'color': info['color'],
                'selection': info['selection']
            }
            for info in self.online_users.values()
        ]
    
    def get_other_users_selections(self, exclude_sid: str) -> List[Dict[str, Any]]:
        """获取其他用户的选中状态"""
        result = []
        for sid, info in self.online_users.items():
            if sid != exclude_sid and info['selection']:
                result.append({
                    'username': info['username'],
                    'color': info['color'],
                    'selection': info['selection']
                })
        return result
    
    def is_empty(self) -> bool:
        return len(self.online_users) == 0


# 全局会话管理：{rel_path: FileSession}
file_sessions: Dict[str, FileSession] = {}

# 记录每个 sid 当前所在的文件 room
user_current_file: Dict[str, str] = {}


def get_or_create_session(filepath: str) -> FileSession:
    """获取或创建文件会话"""
    rel = rel_path(filepath)
    if rel not in file_sessions:
        file_sessions[rel] = FileSession(filepath)
    return file_sessions[rel]


def cleanup_empty_sessions() -> None:
    """清理无人在线的会话，释放内存"""
    empty = [rel for rel, sess in file_sessions.items() if sess.is_empty()]
    for rel in empty:
        del file_sessions[rel]
